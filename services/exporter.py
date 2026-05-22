from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_COLUMNS = [
    "id",
    "campaign_id",
    "name",
    "address",
    "phone",
    "website",
    "google_maps_url",
    "lead_score",
    "priority",
    "estimated_potential",
    "service_fit",
    "suggested_service",
    "possible_pain_points",
    "reasoning_summary",
    "sales_opening_line",
    "email_subject",
    "email_body",
    "decision_maker_name",
    "decision_maker_title",
    "decision_maker_email",
    "decision_maker_phone",
    "linkedin_url",
    "assigned_to",
    "sales_stage",
    "next_follow_up_date",
    "last_contacted_date",
    "notes",
    "source",
    "website_summary",
    "created_at",
    "updated_at",
]


def export_leads_to_excel(leads: list[dict[str, Any]], output_path: Path) -> Path:
    rows = [{col: lead.get(col, "") for col in EXPORT_COLUMNS} for lead in leads]
    df = pd.DataFrame(rows, columns=EXPORT_COLUMNS)
    df.columns = [col.replace("_", " ").title() for col in df.columns]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "LeadRadar Export"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, column in enumerate(ws.columns, start=1):
        max_len = 12
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    wb.save(output_path)
    return output_path
